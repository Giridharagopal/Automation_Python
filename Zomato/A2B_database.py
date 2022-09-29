class A2b_details():
    itempriceA2B={}
    def menu_price(self):
        import openpyxl
        x=openpyxl.load_workbook('C:\\Users\\TGIRIT\\Desktop\\!\\A2B.xlsx')
        z=x["Sheet1"]
        #c=z.cell(row=1,column=1).value
        maxrow=z.max_row
        maxcolumn=z.max_column
        for i in range(1,maxrow+1):
            for j in range(1,maxcolumn+1):
                 q=(z.cell(row=i,column=j).value)
                 self.itempriceA2B["dish"+i]=q
        return self.itempriceA2B

obj=A2b_details()
obj.menu_rice()


